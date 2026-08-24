"""Extraccion de texto de archivos que un agente de Telegram conoce de
memoria (menu, catalogo, lista de precios). El resultado es texto plano
que se agrega siempre al prompt del agente -no hay busqueda dentro del
documento, es contexto fijo, asi que interesa el contenido legible, no el
formato original.
"""

from __future__ import annotations

import csv
import io

from ..core.errors import ValidationError

MAX_FILE_BYTES = 8_000_000  # 8 MB: de sobra para un menu o una lista de precios
MAX_EXTRACTED_CHARS = 20_000  # tope duro por archivo antes de truncar


def extract_text(filename: str, content: bytes) -> str:
    """Nunca devuelve texto vacio silenciosamente sin avisar: si no se
    pudo leer nada, lanza ValidationError -mejor que un agente con un
    archivo "cargado" que en realidad no aporta nada."""
    if len(content) > MAX_FILE_BYTES:
        raise ValidationError(f"El archivo supera el limite de {MAX_FILE_BYTES // 1_000_000} MB.")

    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "pdf":
        text = _extract_pdf(content)
    elif suffix in ("xlsx", "xlsm"):
        text = _extract_xlsx(content)
    elif suffix == "csv":
        text = _extract_csv(content)
    else:
        text = _extract_plain_text(content)

    text = text.strip()
    if not text:
        raise ValidationError("No se pudo extraer texto legible de ese archivo.")
    return text[:MAX_EXTRACTED_CHARS]


def _extract_pdf(content: bytes) -> str:
    from pypdf import PdfReader
    from pypdf.errors import PdfReadError

    try:
        reader = PdfReader(io.BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
    except PdfReadError as error:
        raise ValidationError("Ese PDF no se pudo leer (puede estar danado o protegido).") from error
    return "\n\n".join(page.strip() for page in pages if page.strip())


def _extract_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl no tiene una excepcion propia acotada
        raise ValidationError("Ese Excel no se pudo leer.") from error

    sheets: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(value) for value in row if value is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"[{sheet.title}]\n" + "\n".join(rows))
    return "\n\n".join(sheets)


def _extract_csv(content: bytes) -> str:
    text = _decode(content)
    reader = csv.reader(io.StringIO(text))
    return "\n".join(" | ".join(row) for row in reader if row)


def _extract_plain_text(content: bytes) -> str:
    text = _decode(content)
    if not _looks_like_text(text):
        # latin-1 decodifica CUALQUIER secuencia de bytes sin lanzar -asi
        # que un binario real (una imagen, por ejemplo) "decodifica" igual,
        # solo que como texto ilegible. Sin este chequeo, cualquier archivo
        # con una extension desconocida pasaba como si fuera texto.
        raise ValidationError("Ese archivo no es texto legible (¿es un formato soportado?).")
    return text


def _looks_like_text(text: str) -> bool:
    if not text.strip():
        return False
    printable = sum(1 for char in text if char.isprintable() or char in "\n\r\t")
    return printable / len(text) > 0.9


def _decode(content: bytes) -> str:
    for encoding in ("utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValidationError("Ese archivo no es texto legible (¿es un formato soportado?).")
